"""
Excel parsing utilities for bank statement uploads
Supports both .xlsx (Excel 2007+) and .xls (legacy Excel)
Uses openpyxl for .xlsx and xlrd for .xls
"""
import io
import openpyxl
import xlrd
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from .exceptions import InvalidFileFormatError, ColumnDetectionError
from .csv_parser import _parse_date, _parse_amount, _find_column, _calculate_net_amount


def _detect_header_row_xlsx(sheet, max_search_rows: int = 20) -> Tuple[int, List[str]]:
    """
    Detect which row contains the header in .xlsx file

    Excel files often have:
    - Logo/bank name in first few rows
    - Account info
    - Statement period
    - THEN column headers
    - THEN transaction data

    Args:
        sheet: openpyxl worksheet
        max_search_rows: Maximum rows to search for header

    Returns:
        Tuple[int, List[str]]: (header_row_index, header_values)
                               Row index is 0-based
    """
    max_rows = min(max_search_rows, sheet.max_row)

    for row_idx in range(max_rows):
        # openpyxl is 1-indexed, so add 1
        row = sheet[row_idx + 1]
        row_values = [cell.value for cell in row if cell.value]

        if not row_values:
            continue

        # Check if this row looks like a header
        row_text = ' '.join(str(v).lower() for v in row_values)

        header_keywords = [
            'date', 'description', 'amount', 'transaction',
            'payee', 'merchant', 'debit', 'credit', 'balance',
            'narration', 'withdrawal', 'deposit'
        ]

        if any(keyword in row_text for keyword in header_keywords):
            headers = [str(cell.value).strip() if cell.value else '' for cell in row]
            print(f"Detected header row at index {row_idx}: {headers[:7]}")  # Show first 7 columns
            return row_idx, headers

    # Default to first row if no header detected
    print("No clear header detected, using first row")
    first_row = sheet[1]
    headers = [str(cell.value).strip() if cell.value else '' for cell in first_row]
    return 0, headers


def parse_xlsx(file_content: bytes) -> List[Dict]:
    """
    Parse .xlsx file (Excel 2007+) using openpyxl

    Args:
        file_content: Raw .xlsx file bytes

    Returns:
        List[Dict]: List of transaction dictionaries

    Raises:
        InvalidFileFormatError: If file is corrupted or invalid
        ColumnDetectionError: If required columns not found
    """
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = workbook.active  # Get first/active sheet

        print(f"Processing Excel file with {sheet.max_row} rows, {sheet.max_column} columns")

        # Detect header row
        header_row_idx, header_row = _detect_header_row_xlsx(sheet)

        # Normalize headers
        normalized_headers = {
            i: str(h).lower().strip()
            for i, h in enumerate(header_row)
            if h
        }

        print(f"Normalized headers: {normalized_headers}")

        # Find columns using same logic as PDF parser
        date_col_idx = _find_column_index(
            normalized_headers,
            ['date', 'transaction date', 'posting date', 'trans date', 'trans. date']
        )

        desc_col_idx = _find_column_index(
            normalized_headers,
            ['description', 'narration', 'payee', 'merchant', 'details', 'memo', 'narrative', 'particulars']
        )

        amount_col_idx = _find_column_index(
            normalized_headers,
            ['amount', 'debit', 'withdrawal', 'withdrawalamt.', 'payment',
             'debit amount', 'withdrawals', 'withdrawal amt.', 'withdrawal amt',
             'dr', 'debit amt']
        )

        # Validate columns found
        if date_col_idx is None:
            raise ColumnDetectionError(
                f"Could not identify date column. Found headers: {list(normalized_headers.values())}"
            )

        if desc_col_idx is None:
            raise ColumnDetectionError(
                f"Could not identify description column. Found headers: {list(normalized_headers.values())}"
            )

        if amount_col_idx is None:
            raise ColumnDetectionError(
                f"Could not identify amount column. Found headers: {list(normalized_headers.values())}"
            )

        print(f"Detected columns - Date: {date_col_idx}, Description: {desc_col_idx}, Amount: {amount_col_idx}")

        # Check for credit column (for net amount calculation)
        credit_col_idx = _find_column_index(
            normalized_headers,
            ['credit', 'deposit', 'deposits', 'depositamt.', 'credit amount',
             'deposit amt.', 'deposit amt', 'cr', 'credit amt']
        )

        # Parse transactions (skip header row and any rows before it)
        transactions = []

        for row_idx in range(header_row_idx + 2, sheet.max_row + 1):  # openpyxl is 1-indexed
            row = sheet[row_idx]

            try:
                # Get cell values
                date_cell = row[date_col_idx].value if date_col_idx < len(row) else None
                desc_cell = row[desc_col_idx].value if desc_col_idx < len(row) else None
                amount_cell = row[amount_col_idx].value if amount_col_idx < len(row) else None

                # Parse date
                date = _parse_date(date_cell)

                # Parse description
                description = str(desc_cell).strip() if desc_cell else ''

                # Validate date and description
                if not date:
                    continue
                if not description:
                    continue

                # Handle amount calculation
                transaction_type = 'expense'  # Default

                if credit_col_idx is not None and credit_col_idx < len(row):
                    # Both debit and credit columns exist
                    credit_cell = row[credit_col_idx].value

                    result = _calculate_net_amount(
                        str(amount_cell) if amount_cell else '',
                        str(credit_cell) if credit_cell else ''
                    )

                    if result is None:
                        continue  # Skip empty or balanced rows

                    amount = result['amount']
                    transaction_type = result['type']
                else:
                    # Single amount column only
                    amount = _parse_amount(str(amount_cell)) if amount_cell else None
                    if not amount or amount == 0:
                        continue

                    if amount < 0:
                        amount = abs(amount)
                        transaction_type = 'income'
                    else:
                        transaction_type = 'expense'

                transactions.append({
                    'date': date,
                    'description': description,
                    'amount': amount,
                    'type': transaction_type
                })

            except Exception as e:
                # Skip problematic rows
                print(f"  Warning: Skipping row {row_idx} due to error: {str(e)}")
                continue

        if not transactions:
            raise InvalidFileFormatError("No valid transactions found in Excel file")

        print(f"Successfully parsed {len(transactions)} transaction(s)")
        return transactions

    except openpyxl.utils.exceptions.InvalidFileException as e:
        raise InvalidFileFormatError(f"Invalid or corrupted .xlsx file: {str(e)}")
    except Exception as e:
        if isinstance(e, (InvalidFileFormatError, ColumnDetectionError)):
            raise
        raise InvalidFileFormatError(f"Failed to parse .xlsx file: {str(e)}")


def parse_xls(file_content: bytes) -> List[Dict]:
    """
    Parse legacy .xls file (Excel 97-2003) using xlrd

    Args:
        file_content: Raw .xls file bytes

    Returns:
        List[Dict]: List of transaction dictionaries

    Raises:
        InvalidFileFormatError: If file is corrupted or invalid
        ColumnDetectionError: If required columns not found
    """
    try:
        # Open workbook
        workbook = xlrd.open_workbook(file_contents=file_content)
        sheet = workbook.sheet_by_index(0)  # Get first sheet

        print(f"Processing legacy Excel file with {sheet.nrows} rows, {sheet.ncols} columns")

        # Detect header row (simpler for xlrd)
        header_row_idx = 0
        for row_idx in range(min(20, sheet.nrows)):
            row_values = sheet.row_values(row_idx)
            row_text = ' '.join(str(v).lower() for v in row_values if v)

            header_keywords = [
                'date', 'description', 'amount', 'transaction',
                'payee', 'merchant', 'debit', 'credit', 'narration'
            ]

            if any(keyword in row_text for keyword in header_keywords):
                header_row_idx = row_idx
                print(f"Detected header row at index {row_idx}")
                break

        # Get headers
        headers = sheet.row_values(header_row_idx)
        normalized_headers = {
            i: str(h).lower().strip()
            for i, h in enumerate(headers)
            if h
        }

        print(f"Normalized headers: {normalized_headers}")

        # Find columns
        date_col_idx = _find_column_index(
            normalized_headers,
            ['date', 'transaction date', 'posting date', 'trans date']
        )

        desc_col_idx = _find_column_index(
            normalized_headers,
            ['description', 'narration', 'payee', 'merchant', 'details', 'memo']
        )

        amount_col_idx = _find_column_index(
            normalized_headers,
            ['amount', 'debit', 'withdrawal', 'withdrawalamt.', 'payment']
        )

        # Validate columns
        if date_col_idx is None:
            raise ColumnDetectionError(
                f"Could not identify date column. Found headers: {list(normalized_headers.values())}"
            )

        if desc_col_idx is None:
            raise ColumnDetectionError(
                f"Could not identify description column. Found headers: {list(normalized_headers.values())}"
            )

        if amount_col_idx is None:
            raise ColumnDetectionError(
                f"Could not identify amount column. Found headers: {list(normalized_headers.values())}"
            )

        print(f"Detected columns - Date: {date_col_idx}, Description: {desc_col_idx}, Amount: {amount_col_idx}")

        # Check for credit column
        credit_col_idx = _find_column_index(
            normalized_headers,
            ['credit', 'deposit', 'depositamt.']
        )

        # Parse transactions
        transactions = []

        for row_idx in range(header_row_idx + 1, sheet.nrows):
            try:
                # Get cell values
                date_value = sheet.cell_value(row_idx, date_col_idx)
                desc_value = sheet.cell_value(row_idx, desc_col_idx)
                amount_value = sheet.cell_value(row_idx, amount_col_idx)

                # Handle Excel date numbers (xlrd converts dates to floats)
                if isinstance(date_value, float):
                    date = xlrd.xldate_as_datetime(date_value, workbook.datemode)
                else:
                    date = _parse_date(date_value)

                # Parse description
                description = str(desc_value).strip() if desc_value else ''

                # Validate date and description
                if not date or not description:
                    continue

                # Handle amount calculation
                transaction_type = 'expense'  # Default

                if credit_col_idx is not None:
                    # Both debit and credit columns exist
                    credit_value = sheet.cell_value(row_idx, credit_col_idx)

                    result = _calculate_net_amount(
                        str(amount_value) if amount_value else '',
                        str(credit_value) if credit_value else ''
                    )

                    if result is None:
                        continue  # Skip empty or balanced rows

                    amount = result['amount']
                    transaction_type = result['type']
                else:
                    # Single amount column only
                    amount = _parse_amount(str(amount_value)) if amount_value else None
                    if not amount or amount == 0:
                        continue

                    if amount < 0:
                        amount = abs(amount)
                        transaction_type = 'income'
                    else:
                        transaction_type = 'expense'

                transactions.append({
                    'date': date,
                    'description': description,
                    'amount': amount,
                    'type': transaction_type
                })

            except Exception as e:
                print(f"  Warning: Skipping row {row_idx} due to error: {str(e)}")
                continue

        if not transactions:
            raise InvalidFileFormatError("No valid transactions found in Excel file")

        print(f"Successfully parsed {len(transactions)} transaction(s)")
        return transactions

    except xlrd.biffh.XLRDError as e:
        raise InvalidFileFormatError(f"Invalid or corrupted .xls file: {str(e)}")
    except Exception as e:
        if isinstance(e, (InvalidFileFormatError, ColumnDetectionError)):
            raise
        raise InvalidFileFormatError(f"Failed to parse .xls file: {str(e)}")


def _find_column_index(normalized_headers: Dict[int, str], possible_names: List[str]) -> Optional[int]:
    """
    Find column index by possible column names

    Args:
        normalized_headers: Dict mapping {column_index: normalized_name}
        possible_names: List of possible column names to match

    Returns:
        int: Column index if found
        None: If no match found
    """
    for idx, header in normalized_headers.items():
        if header in possible_names:
            return idx
    return None


def parse_excel_file(file_content: bytes, file_type: str) -> List[Dict]:
    """
    Main entry point for Excel parsing

    Args:
        file_content: Raw Excel file bytes
        file_type: 'excel_xlsx' or 'excel_xls'

    Returns:
        List[Dict]: List of transaction dictionaries with keys:
                   - date: datetime object
                   - description: str
                   - amount: float

    Raises:
        InvalidFileFormatError: If file type not supported or corrupted
        ColumnDetectionError: If required columns not found
    """
    if file_type == 'excel_xlsx':
        return parse_xlsx(file_content)
    elif file_type == 'excel_xls':
        return parse_xls(file_content)
    else:
        raise InvalidFileFormatError(f"Unsupported Excel file type: {file_type}")
